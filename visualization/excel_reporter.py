"""
Excel Report Generator
Creates comprehensive Excel reports with multiple sheets
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any
import logging

logger = logging.getLogger(__name__)


class ExcelReporter:
    """Generate comprehensive Excel reports"""
    
    def __init__(self, output_dir: str = "results/reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Style definitions
        self.header_font = Font(bold=True, size=14, color="FFFFFF")
        self.header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        self.subheader_font = Font(bold=True, size=12)
        self.subheader_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_comprehensive_report(self,
                                     metrics: Any,
                                     cluster_data: Dict,
                                     solar_data: Dict,
                                     economic_data: Dict,
                                     temporal_data: pd.DataFrame = None) -> str:
        """Generate comprehensive Excel report with multiple sheets"""
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # 1. Executive Summary
        self._create_executive_summary(wb, metrics, economic_data)
        
        # 2. Cluster Analysis
        self._create_cluster_analysis(wb, cluster_data)
        
        # 3. Building Details
        self._create_building_details(wb, metrics, solar_data)
        
        # 4. Energy Flows
        self._create_energy_flows(wb, metrics, temporal_data)
        
        # 5. Solar Analysis
        self._create_solar_analysis(wb, solar_data)
        
        # 6. Economic Analysis
        self._create_economic_analysis(wb, economic_data)
        
        # 7. Network Metrics
        self._create_network_metrics(wb, metrics)
        
        # 8. Temporal Analysis
        if temporal_data is not None:
            self._create_temporal_analysis(wb, temporal_data)
        
        # Save workbook
        filename = f"energy_community_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        logger.info(f"Excel report saved to {filepath}")
        return str(filepath)
    
    def _create_executive_summary(self, wb: Workbook, metrics: Any, economic_data: Dict):
        """Create executive summary sheet"""
        
        ws = wb.create_sheet("Executive Summary")
        
        # Title
        ws['A1'] = "Energy Community Executive Summary"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
        ws['A2'].font = Font(italic=True)
        
        # Key metrics section
        row = 4
        ws[f'A{row}'] = "KEY PERFORMANCE INDICATORS"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        # KPI data
        kpis = [
            ("Total Clusters", metrics.num_clusters),
            ("Average Cluster Size", f"{metrics.avg_cluster_size:.1f}"),
            ("Cluster Stability", f"{metrics.cluster_stability:.1%}"),
            ("Self-Sufficiency", f"{metrics.avg_self_sufficiency:.1%}"),
            ("Peak Reduction", f"{metrics.total_peak_reduction:.1%}"),
            ("Cost Savings (Monthly)", f"€{metrics.total_cost_savings_eur:,.0f}"),
            ("CO2 Reduced (Monthly)", f"{metrics.carbon_reduction_tons:.1f} tons"),
            ("Solar Coverage", f"{metrics.solar_coverage_percent:.1%}")
        ]
        
        row = 6
        for label, value in kpis:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Financial summary section
        row += 2
        ws[f'A{row}'] = "FINANCIAL SUMMARY"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 2
        if economic_data:
            financial_items = [
                ("Total Investment Required", f"€{economic_data.get('total_investment', 0):,.0f}"),
                ("Annual Benefits", f"€{economic_data.get('total_annual_benefit', 0):,.0f}"),
                ("Payback Period", f"{economic_data.get('overall_payback_years', 0):.1f} years"),
                ("ROI", f"{economic_data.get('roi_percent', 0):.1f}%"),
                ("NPV (20 years)", f"€{economic_data.get('npv', 0):,.0f}")
            ]
            
            for label, value in financial_items:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
        
        # Add a simple chart
        self._add_summary_chart(ws, metrics, row + 2)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
    
    def _create_cluster_analysis(self, wb: Workbook, cluster_data: Dict):
        """Create cluster analysis sheet"""
        
        ws = wb.create_sheet("Cluster Analysis")
        
        # Headers
        headers = ["Cluster ID", "Size", "LV Group", "Quality Score", "Quality Label",
                  "Self-Sufficiency", "Complementarity", "Peak Reduction", 
                  "Stability", "Energy Shared (kWh)"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        row = 2
        for cluster_id, metrics in cluster_data.items():
            ws.cell(row=row, column=1, value=cluster_id)
            ws.cell(row=row, column=2, value=metrics.get('member_count', 0))
            ws.cell(row=row, column=3, value=metrics.get('lv_group_id', ''))
            ws.cell(row=row, column=4, value=metrics.get('quality_score', 0))
            ws.cell(row=row, column=5, value=metrics.get('quality_label', ''))
            ws.cell(row=row, column=6, value=metrics.get('self_sufficiency_ratio', 0))
            ws.cell(row=row, column=7, value=metrics.get('complementarity_score', 0))
            ws.cell(row=row, column=8, value=metrics.get('peak_reduction_ratio', 0))
            ws.cell(row=row, column=9, value=metrics.get('temporal_stability', 0))
            ws.cell(row=row, column=10, value=metrics.get('total_shared_kwh', 0))
            
            # Apply conditional formatting for quality
            quality_cell = ws.cell(row=row, column=5)
            if quality_cell.value == 'excellent':
                quality_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif quality_cell.value == 'good':
                quality_cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
            elif quality_cell.value == 'poor':
                quality_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            row += 1
        
        # Add borders
        for row in ws.iter_rows(min_row=1, max_row=row-1, min_col=1, max_col=10):
            for cell in row:
                cell.border = self.border
        
        # Adjust column widths
        for col in range(1, 11):
            ws.column_dimensions[chr(64+col)].width = 15
    
    def _create_building_details(self, wb: Workbook, metrics: Any, solar_data: Dict):
        """Create building details sheet"""
        
        ws = wb.create_sheet("Building Details")
        
        # Create sample data if not available
        building_data = []
        
        if 'buildings' in solar_data:
            for building in solar_data['buildings']:
                building_data.append({
                    'Building ID': building.get('id'),
                    'Energy Label': building.get('energy_label'),
                    'Type': building.get('type'),
                    'Annual Demand (kWh)': building.get('demand'),
                    'Has Solar': building.get('has_solar'),
                    'Solar Capacity (kWp)': building.get('solar_capacity', 0),
                    'Cluster ID': building.get('cluster_id'),
                    'LV Group': building.get('lv_group'),
                    'Priority Score': building.get('priority_score', 0)
                })
        else:
            # Create sample data
            for i in range(20):
                building_data.append({
                    'Building ID': f'B{i:03d}',
                    'Energy Label': np.random.choice(['A', 'B', 'C', 'D', 'E', 'F', 'G']),
                    'Type': np.random.choice(['Residential', 'Commercial', 'Industrial']),
                    'Annual Demand (kWh)': np.random.randint(3000, 50000),
                    'Has Solar': np.random.choice([True, False]),
                    'Solar Capacity (kWp)': np.random.uniform(0, 20) if np.random.random() > 0.5 else 0,
                    'Cluster ID': np.random.randint(1, 6),
                    'LV Group': f'LV_{np.random.randint(1, 11):03d}',
                    'Priority Score': np.random.uniform(0, 1)
                })
        
        # Convert to DataFrame and write
        df = pd.DataFrame(building_data)
        
        # Write headers
        for col, header in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Apply conditional formatting to energy labels
        label_col = df.columns.get_loc('Energy Label') + 1
        for row in range(2, len(df) + 2):
            cell = ws.cell(row=row, column=label_col)
            if cell.value in ['A', 'B']:
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif cell.value in ['C', 'D']:
                cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
            elif cell.value in ['E', 'F', 'G']:
                cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    
    def _create_energy_flows(self, wb: Workbook, metrics: Any, temporal_data: pd.DataFrame):
        """Create energy flows sheet"""
        
        ws = wb.create_sheet("Energy Flows")
        
        # Summary section
        ws['A1'] = "ENERGY FLOW SUMMARY"
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:D1')
        
        flow_data = [
            ("Total Demand", f"{metrics.total_demand_mwh:.2f} MWh"),
            ("Total Generation", f"{metrics.total_generation_mwh:.2f} MWh"),
            ("Shared Energy", f"{metrics.total_shared_energy_mwh:.2f} MWh"),
            ("Grid Import", f"{metrics.grid_import_mwh:.2f} MWh"),
            ("Grid Export", f"{metrics.grid_export_mwh:.2f} MWh"),
            ("Line Losses", f"{metrics.line_loss_percent:.1f}%")
        ]
        
        row = 3
        for label, value in flow_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Hourly flows if available
        if temporal_data is not None and not temporal_data.empty:
            row += 2
            ws[f'A{row}'] = "HOURLY ENERGY FLOWS"
            ws[f'A{row}'].font = self.header_font
            ws[f'A{row}'].fill = self.header_fill
            ws.merge_cells(f'A{row}:F{row}')
            
            # Aggregate hourly data
            hourly = temporal_data.groupby('hour').agg({
                'demand': 'mean',
                'generation': 'mean'
            }).round(2)
            
            row += 2
            ws[f'A{row}'] = "Hour"
            ws[f'B{row}'] = "Avg Demand (kW)"
            ws[f'C{row}'] = "Avg Generation (kW)"
            ws[f'D{row}'] = "Net Flow (kW)"
            
            for col in range(1, 5):
                ws.cell(row=row, column=col).font = Font(bold=True)
                ws.cell(row=row, column=col).fill = self.subheader_fill
            
            row += 1
            for hour in range(24):
                ws[f'A{row}'] = hour
                if hour in hourly.index:
                    ws[f'B{row}'] = hourly.loc[hour, 'demand']
                    ws[f'C{row}'] = hourly.loc[hour, 'generation']
                    ws[f'D{row}'] = hourly.loc[hour, 'demand'] - hourly.loc[hour, 'generation']
                row += 1
    
    def _create_solar_analysis(self, wb: Workbook, solar_data: Dict):
        """Create solar analysis sheet"""
        
        ws = wb.create_sheet("Solar Analysis")
        
        # Title
        ws['A1'] = "SOLAR INSTALLATION ANALYSIS"
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:E1')
        
        # Summary metrics
        row = 3
        ws[f'A{row}'] = "Summary Metrics"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        summary_data = [
            ("Total Solar Capacity", f"{solar_data.get('total_capacity', 0):.1f} kWp"),
            ("Number of Installations", solar_data.get('num_installations', 0)),
            ("Average ROI", f"{solar_data.get('avg_roi', 0):.1f} years"),
            ("Total Investment Required", f"€{solar_data.get('total_investment', 0):,.0f}"),
            ("Annual Generation", f"{solar_data.get('annual_generation', 0):,.0f} kWh"),
            ("CO2 Savings", f"{solar_data.get('co2_savings', 0):.1f} tons/year")
        ]
        
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Priority installations
        row += 2
        ws[f'A{row}'] = "Priority Installations"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 1
        headers = ["Rank", "Building ID", "Energy Label", "Roof Area (m²)", 
                  "Capacity (kWp)", "ROI (years)", "Priority Score"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header).font = Font(bold=True)
        
        # Add sample priority buildings
        row += 1
        if 'priority_list' in solar_data:
            for rank, building in enumerate(solar_data['priority_list'][:20], 1):
                ws.cell(row=row, column=1, value=rank)
                ws.cell(row=row, column=2, value=building.get('id'))
                ws.cell(row=row, column=3, value=building.get('energy_label'))
                ws.cell(row=row, column=4, value=building.get('roof_area'))
                ws.cell(row=row, column=5, value=building.get('capacity'))
                ws.cell(row=row, column=6, value=building.get('roi'))
                ws.cell(row=row, column=7, value=building.get('priority_score'))
                row += 1
    
    def _create_economic_analysis(self, wb: Workbook, economic_data: Dict):
        """Create economic analysis sheet"""
        
        ws = wb.create_sheet("Economic Analysis")
        
        # Title
        ws['A1'] = "ECONOMIC ANALYSIS"
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:D1')
        
        # Investment breakdown
        row = 3
        ws[f'A{row}'] = "Investment Requirements"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        investments = economic_data.get('investments', {})
        for item, amount in investments.items():
            ws[f'A{row}'] = item
            ws[f'B{row}'] = f"€{amount:,.0f}"
            row += 1
        
        # Annual benefits
        row += 2
        ws[f'A{row}'] = "Annual Benefits"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        benefits = economic_data.get('annual_benefits', {})
        for item, amount in benefits.items():
            ws[f'A{row}'] = item
            ws[f'B{row}'] = f"€{amount:,.0f}"
            row += 1
        
        # Financial metrics
        row += 2
        ws[f'A{row}'] = "Financial Metrics"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        metrics = [
            ("Payback Period", f"{economic_data.get('payback_years', 0):.1f} years"),
            ("Net Present Value", f"€{economic_data.get('npv', 0):,.0f}"),
            ("Internal Rate of Return", f"{economic_data.get('irr', 0):.1%}"),
            ("Profitability Index", f"{economic_data.get('profitability_index', 0):.2f}")
        ]
        
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
    
    def _create_network_metrics(self, wb: Workbook, metrics: Any):
        """Create network metrics sheet"""
        
        ws = wb.create_sheet("Network Metrics")
        
        # Title
        ws['A1'] = "NETWORK PERFORMANCE METRICS"
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:D1')
        
        # Grid metrics
        row = 3
        grid_metrics = [
            ("Average Voltage Deviation", f"{metrics.avg_voltage_deviation:.2%}"),
            ("Transformer Utilization", f"{metrics.transformer_utilization_percent:.1%}"),
            ("Line Loss Percentage", f"{metrics.line_loss_percent:.2%}"),
            ("Congestion Events (Monthly)", metrics.congestion_events),
            ("Number of LV Groups", metrics.num_lv_groups),
            ("Avg Buildings per LV Group", f"{metrics.avg_buildings_per_lv:.1f}"),
            ("LV Groups with Clusters", metrics.lv_groups_with_clusters)
        ]
        
        for label, value in grid_metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Add utilization chart placeholder
        row += 2
        ws[f'A{row}'] = "Transformer Utilization by Location"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:D{row}')
    
    def _create_temporal_analysis(self, wb: Workbook, temporal_data: pd.DataFrame):
        """Create temporal analysis sheet"""
        
        ws = wb.create_sheet("Temporal Analysis")
        
        # Daily profile
        ws['A1'] = "TEMPORAL PATTERNS"
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:E1')
        
        # Aggregate by hour
        if 'hour' in temporal_data.columns:
            hourly = temporal_data.groupby('hour').agg({
                'demand': ['mean', 'std', 'min', 'max']
            }).round(2)
            
            row = 3
            ws[f'A{row}'] = "Hour"
            ws[f'B{row}'] = "Avg Demand"
            ws[f'C{row}'] = "Std Dev"
            ws[f'D{row}'] = "Min"
            ws[f'E{row}'] = "Max"
            
            for col in range(1, 6):
                ws.cell(row=row, column=col).font = Font(bold=True)
                ws.cell(row=row, column=col).fill = self.subheader_fill
            
            row += 1
            for hour in hourly.index:
                ws[f'A{row}'] = hour
                ws[f'B{row}'] = hourly.loc[hour, ('demand', 'mean')]
                ws[f'C{row}'] = hourly.loc[hour, ('demand', 'std')]
                ws[f'D{row}'] = hourly.loc[hour, ('demand', 'min')]
                ws[f'E{row}'] = hourly.loc[hour, ('demand', 'max')]
                row += 1
    
    def _add_summary_chart(self, ws, metrics: Any, start_row: int):
        """Add a summary chart to the worksheet"""
        
        # Create data for chart
        chart_row = start_row
        ws[f'E{chart_row}'] = "Metric"
        ws[f'F{chart_row}'] = "Value"
        ws[f'E{chart_row}'].font = Font(bold=True)
        ws[f'F{chart_row}'].font = Font(bold=True)
        
        chart_data = [
            ("Self-Sufficiency %", metrics.avg_self_sufficiency * 100),
            ("Peak Reduction %", metrics.total_peak_reduction * 100),
            ("Solar Coverage %", metrics.solar_coverage_percent)
        ]
        
        for i, (label, value) in enumerate(chart_data, 1):
            ws[f'E{chart_row + i}'] = label
            ws[f'F{chart_row + i}'] = value
        
        # Create bar chart
        chart = BarChart()
        chart.title = "Key Performance Metrics"
        chart.y_axis.title = 'Percentage'
        chart.x_axis.title = 'Metrics'
        
        data = Reference(ws, min_col=6, min_row=chart_row, max_row=chart_row + len(chart_data), max_col=6)
        categories = Reference(ws, min_col=5, min_row=chart_row + 1, max_row=chart_row + len(chart_data))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        ws.add_chart(chart, f"E{chart_row + len(chart_data) + 2}")